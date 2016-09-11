# openpyxlから必要な関数（load_workbook）をimport
from openpyxl import load_workbook
import numpy as np  # NumPyもimport

# WorkBookを開く
wb = load_workbook(filename='Sample1.xlsx', read_only=True,
                   data_only=True, use_iterators=True)
# WorkSheetを名前で指定
ws = wb['温度変化']

# あらかじめデータを格納するNumPyのndarrayを作成しておく
Nrow = 11
time_vec = np.zeros(Nrow)
temp_vec = np.zeros(Nrow)

# データの読み出し
for i, row in enumerate(ws.iter_rows(row_offset=1)):
    time_vec[i] = row[0].value
    temp_vec[i] = row[1].value
